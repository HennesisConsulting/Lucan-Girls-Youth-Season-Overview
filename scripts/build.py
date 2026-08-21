#!/usr/bin/env python3
"""
Builds docs/index.html from template/dashboard_template.html, injecting the
JSON found in cell A4 of the "Web Output" sheet of the workbook in /workbook.

Run automatically by .github/workflows/deploy.yml whenever the workbook
changes. Can also be run locally:  python3 scripts/build.py
"""
import glob
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_DIR = ROOT / "workbook"
TEMPLATE_PATH = ROOT / "template" / "dashboard_template.html"
OUTPUT_PATH = ROOT / "docs" / "index.html"
SHEET_NAME = "Web Output"
CELL = "A4"


def find_workbook() -> Path:
    candidates = sorted(glob.glob(str(WORKBOOK_DIR / "*.xlsx")))
    # ignore Excel's temporary lock files like ~$Season_Data_Collection.xlsx
    candidates = [c for c in candidates if not Path(c).name.startswith("~$")]
    if not candidates:
        print(f"ERROR: no .xlsx file found in {WORKBOOK_DIR}", file=sys.stderr)
        sys.exit(1)
    if len(candidates) > 1:
        print(
            f"WARNING: multiple .xlsx files found in {WORKBOOK_DIR}, "
            f"using the first one alphabetically: {candidates[0]}",
            file=sys.stderr,
        )
    return Path(candidates[0])


def recalculate(workbook_path: Path) -> Path:
    """
    Excel only recalculates formulas that were dirty at save time - if the
    file was saved without a full recalc (Ctrl+Alt+F9), cells like the
    Web Output JSON, return-date calculations, etc. can be silently stale
    even though nothing looks wrong. This has bitten real uploads before.
    Force a full recalculation via headless LibreOffice before reading
    anything, so the published site is never built from stale numbers.
    Falls back to reading the workbook as-is if LibreOffice isn't
    available in this environment, rather than failing the whole build.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print("WARNING: LibreOffice not found - reading workbook without forcing "
              "recalculation. Cached formula values could be stale.", file=sys.stderr)
        return workbook_path

    tmpdir = Path(tempfile.mkdtemp(prefix="recalc_"))
    try:
        result = subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(tmpdir), str(workbook_path)],
            capture_output=True, text=True, timeout=120,
        )
        recalced = tmpdir / workbook_path.name
        if result.returncode != 0 or not recalced.exists():
            print(f"WARNING: LibreOffice recalculation failed ({result.returncode}), "
                  f"falling back to reading the workbook as-is:\n{result.stderr}", file=sys.stderr)
            return workbook_path
        print("Recalculated workbook via LibreOffice before reading (avoids stale cached formulas)")
        return recalced
    except subprocess.TimeoutExpired:
        print("WARNING: LibreOffice recalculation timed out - falling back to "
              "reading the workbook as-is.", file=sys.stderr)
        return workbook_path


def extract_json(workbook_path: Path) -> dict:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(
            f"ERROR: workbook has no sheet named '{SHEET_NAME}'. "
            f"Sheets found: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)
    ws = wb[SHEET_NAME]
    raw = ws[CELL].value
    if not raw:
        print(f"ERROR: cell {CELL} on '{SHEET_NAME}' is empty.", file=sys.stderr)
        sys.exit(1)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"ERROR: cell {CELL} does not contain valid JSON: {e}", file=sys.stderr)
        sys.exit(1)
    if "players" not in data or "teamSelection" not in data:
        print(
            "ERROR: parsed JSON is missing expected keys (players, teamSelection). "
            "Is this the right workbook?",
            file=sys.stderr,
        )
        sys.exit(1)
    return data


def redact_sensitive_fields(data: dict) -> dict:
    """
    The public build should show THAT a player is unavailable, but not WHY
    (injury details, medical info, etc. about minors shouldn't sit on a
    public static site with no access control). Strip specific reasons
    from every place they appear in the JSON, not just how the UI displays
    them - the raw JSON is visible to anyone via "view source" on a static
    GitHub Pages site, so redacting only in the UI would not be a real fix.
    """
    GENERIC = "Unavailable"

    for entry in data.get("teamSelection", {}).get("unavailable", []):
        entry["reason"] = GENERIC

    for ns in data.get("teamSelection", {}).get("notSelected", []):
        status = ns.get("status", "")
        if isinstance(status, str) and status.startswith("Unavailable"):
            ns["status"] = GENERIC

    for p in data.get("players", []):
        if p.get("availability") not in (None, "Yes") and p.get("unavailableReason"):
            p["unavailableReason"] = GENERIC

    return data


def build():
    workbook_path = find_workbook()
    print(f"Reading: {workbook_path.name}")

    # Read the original file's own generated timestamp first (before recalculating,
    # since forcing recalculation re-evaluates any NOW()/TODAY()-style formula to
    # the moment this script runs rather than when the workbook was actually saved).
    try:
        original_data = extract_json(workbook_path)
        original_generated = original_data.get("meta", {}).get("generated")
    except SystemExit:
        original_generated = None

    recalced_path = recalculate(workbook_path)
    data = extract_json(recalced_path)
    if original_generated:
        data["meta"]["generated"] = original_generated
    print(f"Parsed OK - {len(data['players'])} players, generated {data['meta'].get('generated')}")

    data = redact_sensitive_fields(data)
    print("Redacted specific unavailability reasons for the public build")

    if not TEMPLATE_PATH.exists():
        print(f"ERROR: template not found at {TEMPLATE_PATH}", file=sys.stderr)
        sys.exit(1)
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    if "__DATA_PLACEHOLDER__" not in template:
        print("ERROR: template is missing the __DATA_PLACEHOLDER__ marker.", file=sys.stderr)
        sys.exit(1)

    json_text = json.dumps(data, ensure_ascii=False)
    output = template.replace("__DATA_PLACEHOLDER__", json_text)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"Wrote: {OUTPUT_PATH} ({len(output):,} bytes)")


if __name__ == "__main__":
    build()
